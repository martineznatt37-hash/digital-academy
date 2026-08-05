"""Genera cronograma Excel con diagrama de Gantt — Digital Academy."""

from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT = r"C:\Users\marti\OneDrive\Desktop\digital-academy\docs\Cronograma_Digital_Academy.xlsx"
PROJECT_START = date(2026, 3, 29)
PROJECT_END = date(2026, 7, 29)

ACTIVITIES = [
    ("Análisis de requerimientos", "Fase 1 - Planificación", date(2026, 3, 29), date(2026, 4, 4), 100),
    ("Definición de niveles educativos", "Fase 1 - Planificación", date(2026, 4, 5), date(2026, 4, 8), 100),
    ("Arquitectura del sistema", "Fase 1 - Planificación", date(2026, 4, 9), date(2026, 4, 14), 100),
    ("Wireframes y diseño UI/UX", "Fase 1 - Planificación", date(2026, 4, 15), date(2026, 4, 21), 100),
    ("Stack tecnológico y estructura", "Fase 1 - Planificación", date(2026, 4, 22), date(2026, 4, 28), 100),
    ("Servidor Express y API REST", "Fase 2 - Backend", date(2026, 4, 29), date(2026, 5, 3), 100),
    ("Base de datos SQLite", "Fase 2 - Backend", date(2026, 5, 4), date(2026, 5, 8), 100),
    ("Registro y login con JWT", "Fase 2 - Backend", date(2026, 5, 9), date(2026, 5, 16), 100),
    ("Páginas de autenticación", "Fase 2 - Backend", date(2026, 5, 17), date(2026, 5, 20), 100),
    ("Catálogo de cursos con contenido", "Fase 2 - Backend", date(2026, 5, 21), date(2026, 5, 28), 100),
    ("Perfil de usuario en tiempo real", "Fase 3 - Avanzado", date(2026, 5, 29), date(2026, 6, 4), 100),
    ("Registro de horas de estudio", "Fase 3 - Avanzado", date(2026, 6, 5), date(2026, 6, 8), 100),
    ("Gráficos de progreso", "Fase 3 - Avanzado", date(2026, 6, 9), date(2026, 6, 12), 100),
    ("Sistema de asesorías", "Fase 3 - Avanzado", date(2026, 6, 13), date(2026, 6, 17), 100),
    ("Panel de maestros", "Fase 3 - Avanzado", date(2026, 6, 18), date(2026, 6, 20), 100),
    ("Chat con asistente IA", "Fase 3 - Avanzado", date(2026, 6, 21), date(2026, 6, 28), 100),
    ("Pruebas de integración", "Fase 4 - Cierre", date(2026, 6, 29), date(2026, 7, 7), 100),
    ("Corrección de errores", "Fase 4 - Cierre", date(2026, 7, 8), date(2026, 7, 12), 100),
    ("Diseño responsive", "Fase 4 - Cierre", date(2026, 7, 13), date(2026, 7, 15), 100),
    ("Documentación y presentación", "Fase 4 - Cierre", date(2026, 7, 16), date(2026, 7, 22), 100),
    ("Revisión final", "Fase 4 - Cierre", date(2026, 7, 26), date(2026, 7, 28), 100),
    ("ENTREGA FINAL", "Fase 4 - Cierre", date(2026, 7, 29), date(2026, 7, 29), 100),
]

PHASE_COLORS = {
    "Fase 1 - Planificación": "4472C4",
    "Fase 2 - Backend": "2E7D32",
    "Fase 3 - Avanzado": "ED7D31",
    "Fase 4 - Cierre": "7030A0",
}

HEADER_FILL = PatternFill("solid", fgColor="1A5F4A")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color="1A5F4A")
SUB_FONT = Font(size=11, color="444444")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def days_between(start, end):
    return (end - start).days + 1


def offset_days(start):
    return (start - PROJECT_START).days


def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def build_summary_sheet(wb):
    ws = wb.active
    ws.title = "Resumen"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 36

    ws["A1"] = "CRONOGRAMA DEL PROYECTO"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Digital Academy — Plataforma Web Educativa"
    ws["A2"].font = SUB_FONT

    info = [
        ("", ""),
        ("Proyecto", "Digital Academy"),
        ("Duración", "4 meses (122 días)"),
        ("Fecha de inicio", "29 de marzo de 2026"),
        ("Entrega final", "29 de julio de 2026"),
        ("Avance total", "100%"),
        ("", ""),
        ("Fase 1", "Planificación y diseño (29 mar – 28 abr) — 100%"),
        ("Fase 2", "Backend, auth y cursos (29 abr – 28 may) — 100%"),
        ("Fase 3", "Perfil, asesorías e IA (29 may – 28 jun) — 100%"),
        ("Fase 4", "Pruebas, docs y entrega (29 jun – 29 jul) — 100%"),
    ]

    for i, (a, b) in enumerate(info, start=4):
        ws.cell(row=i, column=1, value=a).font = Font(bold=bool(a))
        ws.cell(row=i, column=2, value=b)

    ws["A16"] = "Módulos completados"
    ws["A16"].font = Font(bold=True, size=12, color="1A5F4A")

    modules = [
        "Autenticación (login/registro)", "Perfil con horas en tiempo real",
        "9 cursos con contenido", "Asesorías y panel de maestros",
        "Chat IA educativo", "Gamificación y diseño responsive",
        "Documentación y presentación", "Servidor Node.js funcional",
    ]
    row = 17
    for mod in modules:
        ws.cell(row=row, column=1, value=f"✅ {mod}")
        ws.cell(row=row, column=2, value="100%")
        row += 1


def build_cronograma_sheet(wb):
    ws = wb.create_sheet("Cronograma")
    headers = [
        "#", "Actividad", "Fase", "Inicio", "Fin",
        "Duración (días)", "Desfase (días)", "Avance %",
    ]
    widths = [5, 38, 22, 14, 14, 16, 16, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:H1")
    ws["A1"] = "CRONOGRAMA DETALLADO — DIGITAL ACADEMY"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    for col, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, len(headers))

    for idx, (name, phase, start, end, progress) in enumerate(ACTIVITIES, start=1):
        row = idx + 3
        duration = days_between(start, end)
        offset = offset_days(start)
        values = [idx, name, phase, start, end, duration, offset, progress]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = BORDER
            cell.alignment = LEFT if col in (2, 3) else CENTER
            if col in (4, 5):
                cell.number_format = "DD/MM/YYYY"
            if col == 8:
                cell.fill = PatternFill("solid", fgColor="C6EFCE")
                cell.font = Font(bold=True, color="006100")
            if col == 3:
                color = PHASE_COLORS.get(phase, "808080")
                cell.fill = PatternFill("solid", fgColor=color)
                cell.font = Font(color="FFFFFF", bold=True)

    last_row = len(ACTIVITIES) + 3
    ws.auto_filter.ref = f"A3:H{last_row}"
    ws.freeze_panes = "A4"
    return ws, last_row


def build_gantt_sheet(wb, last_row):
    ws = wb.create_sheet("Diagrama Gantt")

    headers = ["Actividad", "Fase", "Desfase", "Duración"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10

    cronograma = wb["Cronograma"]
    for i in range(len(ACTIVITIES)):
        src = i + 4
        dst = i + 2
        ws.cell(row=dst, column=1, value=cronograma.cell(row=src, column=2).value)
        ws.cell(row=dst, column=2, value=cronograma.cell(row=src, column=3).value)
        ws.cell(row=dst, column=3, value=cronograma.cell(row=src, column=7).value)
        ws.cell(row=dst, column=4, value=cronograma.cell(row=src, column=6).value)
        for col in range(1, 5):
            ws.cell(row=dst, column=col).border = BORDER

    # Invisible offset + visible duration = stacked horizontal bar chart
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = "Diagrama de Gantt — Digital Academy (mar–jul 2026)"
    chart.y_axis.title = "Actividades"
    chart.x_axis.title = "Días desde el 29 de marzo de 2026"
    chart.height = 18
    chart.width = 28
    chart.legend = None

    data_offset = Reference(ws, min_col=3, min_row=1, max_row=len(ACTIVITIES) + 1)
    data_duration = Reference(ws, min_col=4, min_row=1, max_row=len(ACTIVITIES) + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(ACTIVITIES) + 1)

    chart.add_data(data_offset, titles_from_data=True)
    chart.add_data(data_duration, titles_from_data=True)
    chart.set_categories(cats)

    # First series (Desfase) — no fill
    chart.series[0].graphicalProperties.solidFill = "FFFFFF"
    chart.series[0].graphicalProperties.line.solidFill = "FFFFFF"

    # Second series (Duración) — green bars
    chart.series[1].graphicalProperties.solidFill = "2E7D32"

    ws.add_chart(chart, "F2")

    # Timeline labels
    ws["F1"] = "Línea de tiempo"
    ws["F1"].font = Font(bold=True, color="1A5F4A")
    milestones = [
        ("Inicio", "29/03/2026"),
        ("Fin Fase 1", "28/04/2026"),
        ("Fin Fase 2", "28/05/2026"),
        ("Fin Fase 3", "28/06/2026"),
        ("ENTREGA FINAL", "29/07/2026"),
    ]
    row = 2
    for label, dt in milestones:
        ws.cell(row=row, column=6, value=label).font = Font(bold=True)
        ws.cell(row=row, column=7, value=dt)
        row += 1

    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 14


def build_fases_sheet(wb):
    ws = wb.create_sheet("Avance por Fase")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14

    ws["A1"] = "AVANCE POR FASE"
    ws["A1"].font = TITLE_FONT

    headers = ["Fase", "Periodo", "Actividades", "Avance %"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, 4)

    phases = [
        ("Fase 1 — Planificación", "29 mar – 28 abr 2026", 5, 100),
        ("Fase 2 — Backend y cursos", "29 abr – 28 may 2026", 5, 100),
        ("Fase 3 — Funciones avanzadas", "29 may – 28 jun 2026", 6, 100),
        ("Fase 4 — Cierre y entrega", "29 jun – 29 jul 2026", 6, 100),
        ("TOTAL PROYECTO", "4 meses", 22, 100),
    ]

    colors = ["4472C4", "2E7D32", "ED7D31", "7030A0", "1A5F4A"]
    for i, (phase, period, acts, pct) in enumerate(phases):
        row = i + 4
        ws.cell(row=row, column=1, value=phase)
        ws.cell(row=row, column=2, value=period)
        ws.cell(row=row, column=3, value=acts)
        ws.cell(row=row, column=4, value=pct)
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            cell.alignment = CENTER if col > 1 else LEFT
            if i < 4:
                cell.fill = PatternFill("solid", fgColor=colors[i])
                if col <= 2:
                    cell.font = Font(color="FFFFFF", bold=True)
            else:
                cell.fill = PatternFill("solid", fgColor="C6EFCE")
                cell.font = Font(bold=True, color="006100")

    # Bar chart for phase progress
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Avance por Fase (%)"
    chart.y_axis.title = "Porcentaje"
    chart.height = 12
    chart.width = 18

    data = Reference(ws, min_col=4, min_row=3, max_row=7)
    cats = Reference(ws, min_col=1, min_row=4, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = "1A5F4A"
    ws.add_chart(chart, "F3")


def main():
    wb = Workbook()
    build_summary_sheet(wb)
    _, last_row = build_cronograma_sheet(wb)
    build_gantt_sheet(wb, last_row)
    build_fases_sheet(wb)
    wb.save(OUTPUT)
    print(f"Archivo creado: {OUTPUT}")


if __name__ == "__main__":
    main()
